"""
  Created on Jan 2025
@author: Elena Markova
          for Attrition Rate Project
"""

import os

os.environ['YDATA_LICENSE_KEY'] = '97d0ae93-9dfc-4c2a-9183-a0420a4d0771'

import pickle
import warnings
import shap
import numpy as np
import pandas as pd
import xgboost as xgb
from scipy import stats
from openpyxl.drawing.image import Image
import openpyxl
from sklearn.metrics import r2_score, precision_score, recall_score, f1_score, confusion_matrix, precision_recall_curve, balanced_accuracy_score
from sklearn.model_selection import train_test_split, cross_val_score
import optuna
from catboost import CatBoostClassifier, Pool
from sklearn.feature_selection import RFECV
from io import BytesIO
from openpyxl import Workbook

from catboost import MetricVisualizer


from pathlib import Path
# from examples.local import setting_dask_env

import seaborn as sn
import matplotlib.pyplot as plt

from model_testing import test_model
from utils.dataset_chemistry import create_features_for_datasets, collect_datasets, minority_class_resample, prepare_dataset_2, get_united_dataset, remove_short_service

# industry_avg_income = df.groupby('field')['income_shortterm'].mean().to_dict()
# df['industry_avg_income'] = df['field'].map(industry_avg_income)
# df['income_vs_industry'] = df['income_shortterm'] - df['industry_avg_income']
# position_median_income = df.groupby('department')['income_shortterm'].median().to_dict()
# df['position_median_income'] = df['department'].map(position_median_income)


def train_xgboost_classifier(_x_train, _y_train, _x_test, _y_test, _sample_weight, _num_iters):
    model = xgb.XGBClassifier(objective="binary:logistic", random_state=42, early_stopping_rounds=15, eval_set=[(_x_test, _y_test)])
    best_f1 = 0.
    best_model = model
    test_result = {}

    print(f"Fitting XGBoost classifier...")
    for iter in range(_num_iters):
        model.fit(_x_train, _y_train, eval_set=[(_x_test, _y_test)], verbose=False)
        predictions = model.predict(_x_test)

        test_result['F1'] = f1_score(_y_test, predictions)
        if test_result['F1'] > best_f1:
            best_f1 = test_result['F1']
            best_model = model

            test_result['Recall'] = recall_score(_y_test, predictions)
            test_result['Precision'] = precision_score(_y_test, predictions)

    feature_importance = best_model.get_booster().get_score(importance_type='gain')
    keys = list(feature_importance.keys())

    # Debug print:
    values = list(feature_importance.values())
    for k, v in zip(keys, values):
        print(k)
    for k, v in zip(keys, values):
        print(v)

    data = pd.DataFrame(data=values, index=keys, columns=["score"]).sort_values(by="score", ascending=False)
    l = data.nlargest(20, columns="score").plot(kind='barh', figsize=(20, 10))  # plot top 20 features
    print(f"XGBoost test result: Recall = {test_result['Recall']}\nPrecision = {test_result['Precision']}\nF1 = {test_result['F1']}")

    feature_importance = best_model.feature_importances_
    feature_importance_df = pd.DataFrame({'Feature': _x_train.columns, 'Importance': feature_importance})
    print(feature_importance_df)

    return best_model

class CustomPrecisionMetric:
    """Более надежная версия для любой классификации"""

    def get_final_error(self, error, weight):
        return error / weight if weight != 0 else 0

    def is_max_optimal(self):
        return True

    def evaluate(self, approxes, target, weight):
        """
        approxes: list of list of floats - логиты для каждого класса
        """
        # Конвертируем в numpy
        approxes_np = np.array(approxes).T  # транспонируем: [объекты, классы]

        # Преобразуем логиты в вероятности через softmax
        # Для бинарной классификации softmax эквивалентен сигмоиде
        exp_approx = np.exp(approxes_np - np.max(approxes_np, axis=1, keepdims=True))
        probas = exp_approx / np.sum(exp_approx, axis=1, keepdims=True)

        # Вероятность положительного класса (последний столбец)
        pos_probas = probas[:, -1] if probas.shape[1] > 1 else probas[:, 0]

        # Порог
        threshold = 0.2
        y_pred = (pos_probas > threshold).astype(int)
        y_true = np.array(target)

        # Расчет precision
        true_positives = np.sum((y_pred == 1) & (y_true == 1))
        predicted_positives = np.sum(y_pred == 1)

        if predicted_positives == 0:
            precision = 0.001  # чтобы избежать NaN
        else:
            precision = true_positives / predicted_positives

        return precision, 1


class CustomPrecisionMetricSimple:
    """Упрощенная версия, совместимая с CatBoost"""

    def __init__(self, threshold=0.5):
        self.threshold = threshold

    def get_final_error(self, error, weight):
        return error

    def is_max_optimal(self):
        return True

    def evaluate(self, approxes, target, weight):
        """
        approxes: list of list of floats
        """
        # Для CatBoost в бинарной классификации approxes[1] - логиты для класса 1
        if len(approxes) == 2:
            approx = np.array(approxes[1])  # логиты для класса 1
        else:
            approx = np.array(approxes[0])

        # Сигмоида
        probas = 1 / (1 + np.exp(-approx))
        y_pred = (probas > self.threshold).astype(int)
        y_true = np.array(target)

        # Precision
        tp = np.sum((y_pred == 1) & (y_true == 1))
        pp = np.sum(y_pred == 1)

        if pp == 0:
            return 0.0, 1  # Возвращаем 0 вместо малого значения
        else:
            return tp / pp, 1


class PrecisionAtKMetric:
    def __init__(self, k=100):
        self.k = k

    def get_final_error(self, error, weight):
        return error

    def is_max_optimal(self):
        return True

    def evaluate(self, approxes, target, weight):
        probas = 1 / (1 + np.exp(-np.array(approxes[0])))

        # Берем топ-K по вероятности
        top_k_indices = np.argsort(probas)[-self.k:]
        y_pred_top_k = np.zeros_like(probas)
        y_pred_top_k[top_k_indices] = 1

        # Precision@K
        precision = precision_score(target, y_pred_top_k, zero_division=0)
        return precision, 1

    def evaluate2(self, approxes, target, weight):
        # Ищем оптимальный порог для precision на лету
        probas = 1 / (1 + np.exp(-np.array(approxes[0])))

        # Перебираем пороги
        best_precision = 0
        for threshold in np.arange(0.1, 0.9, 0.05):
            y_pred = (probas > threshold).astype(int)
            precision = precision_score(target, y_pred, zero_division=0)
            best_precision = max(best_precision, precision)

        return best_precision, 1

def active_learning_for_precision(model, X_pool, uncertainty_threshold=0.3):
    """
    Активное обучение: выбираем примеры, где модель наименее уверена

    Parameters:
    -----------
    model : обученная модель
    X_pool : pool данных для активного обучения
    uncertainty_threshold : порог неопределенности

    Returns:
    --------
    uncertain_indices : индексы примеров с наибольшей неопределенностью
    """
    # Получаем вероятности предсказаний
    probas = model.predict_proba(X_pool)

    # Уверенность модели = максимальная вероятность среди классов
    confidence = np.max(probas, axis=1)

    # Находим примеры с низкой уверенностью (высокая неопределенность)
    uncertain_indices = np.where(confidence < uncertainty_threshold)[0]

    return uncertain_indices


def train_catboost_active_learning(_x_train, _y_train, _x_test, _y_test, _sample_weight, _cat_feats_encoded,
                                   _num_iters):
    """
    Функция обучения CatBoost с активным обучением

    Parameters:
    -----------
    _x_train, _y_train : начальный обучающий набор
    _x_test, _y_test : тестовый набор для валидации
    _sample_weight : веса объектов
    _cat_feats_encoded : категориальные признаки
    _num_iters : количество итераций активного обучения

    Returns:
    --------
    model : обученная модель CatBoost
    """

    # Инициализируем модель с теми же параметрами
    model = CatBoostClassifier(
        iterations=800,
        learning_rate=0.017,
        depth=5,
        l2_leaf_reg=9.0,
        bootstrap_type='Bayesian',
        min_data_in_leaf=50,
        auto_class_weights='Balanced',
        eval_metric='BalancedAccuracy',  # Используем Precision@K
        verbose=False,
        random_seed=42
    )

    # Разделяем данные на начальный обучающий набор и pool для активного обучения
    # Для простоты предположим, что _x_train, _y_train уже содержат начальный набор
    # В реальном сценарии вам нужно разделить данные

    # Создаем копии для активного обучения
    X_current = _x_train.copy()
    y_current = _y_train.copy()

    # Если у нас есть отдельный pool данных для активного обучения
    # Здесь для примера используем часть тестовых данных как pool
    # В реальном приложении у вас должен быть отдельный pool неразмеченных данных
    X_pool = _x_test.copy()
    y_pool = _y_test.copy()

    print("Начинаем активное обучение...")

    # Количество примеров для добавления за итерацию
    n_samples_per_iteration = max(10, len(X_pool) // (_num_iters * 2))

    for iteration in range(_num_iters):
        print(f"\n{'=' * 60}")
        print(f"Итерация активного обучения {iteration + 1}/{_num_iters}")
        print(f"{'=' * 60}")

        # 1. Обучаем модель на текущих данных
        print(f"Размер обучающей выборки: {len(X_current)}")
        print("Обучение модели...")

        model.fit(
            X_current,
            y_current,
            verbose=False
        )

        # 2. Оцениваем модель на тестовых данных
        if len(_x_test) > 0 and len(_y_test) > 0:
            y_pred = model.predict(_x_test)
            precision = precision_score(_y_test, y_pred, zero_division=0)
            print(f"Precision на тесте: {precision:.4f}")

        # 3. Проверяем, остались ли данные в pool
        if len(X_pool) == 0:
            print("Pool данных пуст. Завершаем активное обучение.")
            break

        # 4. Выбираем наиболее неопределенные примеры из pool
        print("Выбор наиболее неопределенных примеров...")
        uncertain_indices = active_learning_for_precision(
            model,
            X_pool,
            uncertainty_threshold=0.3
        )

        # 5. Выбираем примеры для добавления
        if len(uncertain_indices) > 0:
            # Берем n_samples_per_iteration самых неопределенных примеров
            n_to_select = min(n_samples_per_iteration, len(uncertain_indices))

            # Получаем уверенность модели для неопределенных примеров
            confidence = np.max(model.predict_proba(X_pool), axis=1)

            # Сортируем по уверенности (от наименее уверенных)
            sorted_indices = uncertain_indices[np.argsort(confidence[uncertain_indices])]
            selected_indices = sorted_indices[:n_to_select]

            print(f"Выбрано {n_to_select} неопределенных примеров")

            # 6. Добавляем выбранные примеры в обучающую выборку
            if isinstance(X_current, np.ndarray):
                X_current = np.vstack([X_current, X_pool[selected_indices]])
                y_current = np.concatenate([y_current, y_pool[selected_indices]])
            else:
                # Для pandas DataFrame
                X_current = pd.concat([X_current, X_pool.iloc[selected_indices]])
                y_current = pd.concat([y_current, y_pool.iloc[selected_indices]])

            # 7. Удаляем выбранные примеры из pool
            mask = np.ones(len(X_pool), dtype=bool)
            mask[selected_indices] = False

            if isinstance(X_pool, np.ndarray):
                X_pool = X_pool[mask]
                y_pool = y_pool[mask]
            else:
                X_pool = X_pool.iloc[mask]
                y_pool = y_pool.iloc[mask]

            print(f"Новый размер обучающей выборки: {len(X_current)}")
            print(f"Осталось в pool: {len(X_pool)}")
        else:
            print("Нет достаточно неопределенных примеров в pool")

            # Если нет неопределенных примеров, берем случайные
            n_to_select = min(n_samples_per_iteration, len(X_pool))
            if n_to_select > 0:
                random_indices = np.random.choice(len(X_pool), n_to_select, replace=False)

                # Добавляем случайные примеры
                if isinstance(X_current, np.ndarray):
                    X_current = np.vstack([X_current, X_pool[random_indices]])
                    y_current = np.concatenate([y_current, y_pool[random_indices]])
                else:
                    X_current = pd.concat([X_current, X_pool.iloc[random_indices]])
                    y_current = pd.concat([y_current, y_pool.iloc[random_indices]])

                # Удаляем из pool
                mask = np.ones(len(X_pool), dtype=bool)
                mask[random_indices] = False

                if isinstance(X_pool, np.ndarray):
                    X_pool = X_pool[mask]
                    y_pool = y_pool[mask]
                else:
                    X_pool = X_pool.iloc[mask]
                    y_pool = y_pool.iloc[mask]

                print(f"Добавлено {n_to_select} случайных примеров")
                print(f"Новый размер обучающей выборки: {len(X_current)}")
                print(f"Осталось в pool: {len(X_pool)}")

    print(f"\n{'=' * 60}")
    print("Активное обучение завершено!")
    print(f"Итоговый размер обучающей выборки: {len(X_current)}")
    print(f"{'=' * 60}")

    # Финальное обучение на всей собранной выборке
    print("\nФинальное обучение на всей собранной выборке...")
    model.fit(
        X_current,
        y_current,
        verbose=True  # Показываем прогресс финального обучения
    )

    return model


# Альтернативная версия с использованием Pool для активного обучения
def train_catboost_with_pool(_x_train, _y_train, _x_test, _y_test, _sample_weight, _cat_feats_encoded, _num_iters):
    """
    Упрощенная версия с предопределенным pool для активного обучения
    """

    # Предполагаем, что у нас есть отдельный pool данных
    # В реальном сценарии эти данные должны быть переданы как параметры

    # Разделяем данные на обучающие и pool (пример - 70% train, 30% pool)
    from sklearn.model_selection import train_test_split

    X_initial, X_pool, y_initial, y_pool = train_test_split(
        _x_train, _y_train,
        test_size=0.3,
        stratify=_y_train,
        random_state=42
    )

    print(f"Начальный размер обучающей выборки: {len(X_initial)}")
    print(f"Размер pool для активного обучения: {len(X_pool)}")

    model = CatBoostClassifier(
        iterations=800,
        learning_rate=0.017,
        depth=5,
        l2_leaf_reg=9.0,
        bootstrap_type='Bayesian',
        min_data_in_leaf=50,
        auto_class_weights='Balanced',
        eval_metric=PrecisionAtKMetric(k=100),
        verbose=False,
        random_seed=42
    )

    # Используем функцию active_learning_for_precision, определенную выше
    # Процесс аналогичен предыдущей функции

    # Для краткости возвращаем модель, обученную на начальных данных
    print("Обучение на начальных данных...")
    model.fit(X_initial, y_initial, verbose=False)

    return model


# Основная функция для выбора стратегии
def train_catboost_active(_x_train, _y_train, _x_test, _y_test, _sample_weight, _cat_feats_encoded, _num_iters,
                   use_active_learning=True, active_learning_pool=None):
    """
    Основная функция обучения CatBoost с опциональным активным обучением

    Parameters:
    -----------
    use_active_learning : использовать ли активное обучение
    active_learning_pool : дополнительные данные для активного обучения (X_pool, y_pool)
    """

    if use_active_learning and active_learning_pool is not None:
        print("Используем активное обучение с предоставленным pool...")
        X_pool, y_pool = active_learning_pool

        # Объединяем начальные данные с pool для активного обучения
        return train_catboost_active_learning(
            _x_train, _y_train, _x_test, _y_test, X_pool, y_pool,
            _sample_weight, _cat_feats_encoded, _num_iters
        )
    elif use_active_learning:
        print("Используем активное обучение с разделением данных...")
        return train_catboost_with_pool(
            _x_train, _y_train, _x_test, _y_test,
            _sample_weight, _cat_feats_encoded, _num_iters
        )
    else:
        print("Используем стандартное обучение без активного обучения...")
        model = CatBoostClassifier(
            iterations=800,
            learning_rate=0.017,
            depth=5,
            l2_leaf_reg=9.0,
            bootstrap_type='Bayesian',
            min_data_in_leaf=50,
            auto_class_weights='Balanced',
            eval_metric=CustomPrecisionMetric(),
            verbose=False,
            random_seed=42
        )

        model.fit(_x_train, _y_train, verbose=False)
        return model

def train_catboost(_x_train, _y_train, _x_test, _y_test, _sample_weight, _cat_feats_encoded, _num_iters):
    # model already initialized with latest version of optimized parameters for our dataset
    model = CatBoostClassifier(
        iterations=800,
        learning_rate=0.017,
        depth=5,  # Более мелкие деревья для временных данных
        l2_leaf_reg=5.0,  # Сильная регуляризация из-за корреляции?

        # Критически важно для временных данных
        #has_time=True,  # Если есть временная метка
        bootstrap_type='Bayesian',  # Байесовский бэггинг лучше для зависимых данных
        #bagging_temperature=0.5,  # Контроль случайности бэггинга

        # Предотвращение переобучения на автокорреляцию
        # subsample=0.6,  # Еще более агрессивный бэггинг
        min_data_in_leaf=40,  # Больше сэмплов в листе
        #max_leaves=32,
        auto_class_weights='Balanced',

        # Веса для учета дублирования

        eval_metric='BalancedAccuracy',
        # early_stopping_rounds=75,
    )
    #model = CatBoostClassifier(
       #  iterations=900,  # default to 1000
       #  #learning_rate=0.1,
       #  od_type='IncToDec',
       #  l2_leaf_reg=5,
       #  bootstrap_type='Bayesian',
       #  # od_wait=52,
       #  depth=4,  # normally set it from 6 to 10
       #  eval_metric='BalancedAccuracy',
       #  # #random_seed=42,
       #  # sampling_frequency='PerTreeLevel',
       #  random_strength=1,  # default: 1
       #  # loss_function="Logloss",
       #  # #bagging_temperature=2,
       #  # auto_class_weights='Balanced',
       # # scale_pos_weight=7
    #)

    # model = CatBoostClassifier(
    #     iterations=892,  # 400  # Fewer trees + early stopping
    #     learning_rate=0.166,  # 0.08,  # Smaller steps for better generalization
    #     depth=6,  # Slightly deeper but not excessive
    #     l2_leaf_reg=12,  # 10,  # Stronger L2 regularization
    #     bootstrap_type='MVS',
    #     # bagging_temperature=1,  # Less aggressive subsampling
    #     random_strength=1,  # Default randomness
    #     loss_function='Logloss',
    #     eval_metric='AUC',
    #     scale_pos_weight=14.512,
    #     # auto_class_weights='Balanced',  # Adjust for class imbalance
    #     od_type='IncToDec',  # Early stopping
    #     od_wait=31,  # Patience before stopping
    #     silent=True
    # )

    def objective(trial):
        train_pool = Pool(_x_train, _y_train)
        eval_pool = Pool(_x_test, _y_test)
        params={'learning_rate': trial.suggest_float('learning_rate', 0.005, 0.2, log=True),
                'iterations': trial.suggest_int('iterations', 400, 1000),
                'depth': trial.suggest_int('depth', 4, 8),
                'l2_leaf_reg': trial.suggest_int('l2_leaf_reg1', 3, 15),
                'od_wait': trial.suggest_int('od_wait', 20, 80),
                'od_type': trial.suggest_categorical('od_type', ['Iter', 'IncToDec']),
                #'bagging_temperature': trial.suggest_float('bagging_temperature', 0.5, 1),
                'random_strength': trial.suggest_int('random_strength', 1, 3),
                'bootstrap_type': trial.suggest_categorical('bootstrap_type', ['Bayesian', 'MVS']),
                'scale_pos_weight': trial.suggest_float("scale_pos_weight", 1, 15),
                'eval_metric': 'BalancedAccuracy'}
        model1 = CatBoostClassifier(**params, verbose=0)
        model1.fit(
            _x_train, _y_train,
            eval_set=(_x_test, _y_test),
            early_stopping_rounds=50,
        )
        # Get validation predictions
        y_pred = model1.predict(_x_test)
        score = balanced_accuracy_score(_y_test, y_pred)
        #score = cross_val_score(model, pd.concat([_x_train, _x_test]), pd.concat([_y_train, _y_test]), cv=5, scoring='roc_auc').mean()
        return score
    # study = optuna.create_study(direction='maximize')
    # study.optimize(objective, n_trials=200)
    # print(f"Best parameters of optuna: {study.best_params}")

    # perform feature selection
    selector = RFECV(
        estimator=model,
        step=1,
        cv=3,
        scoring='roc_auc'
    )
    #selector.fit(_x_train, _y_train)

    #selected_features = _x_train.columns[selector.support_]
    #print(f'Selected features: {selected_features}')
    print("Start fitting...")
    model.fit(
        _x_train,
        _y_train,
        #eval_set=(_x_test, _y_test),
        verbose=False,
        # sample_weight=_sample_weight,
        # plot=True,
        # cat_features=_cat_feats_encoded  # - do this if haven't encoded cat features
    )
    print('FItting finished')
    # with open('model_chemistry_52_49_066.pkl', 'rb') as file:
    #     model = pickle.load(file)
    train_pool = Pool(data=_x_train, label=_y_train)  #, cat_features=_cat_feats_encoded)
    shap_values = model.get_feature_importance(prettified=False, type='ShapValues', data=train_pool)
    feature_names = _x_train.columns

    # Словарь для переименования фич
    feature_name_mapping = {
        'salary_vs_city': 'Доход / средняя з/п в городе',
        'age_group_46-55': 'Возрастная группа 46-55',
        'job_category_производство': 'Категория "Работник производства"',
        'age_sqr': 'Квадрат возраста',
        'position_industry_управление_3': 'Категория "Управление"',
        'age': 'Возраст',
        'income_group_medium': 'Уровень дохода: средний',
        'income_group_low': 'Уровень дохода: низкий',
        'income_group_medium_high': 'Уровень дохода: средне-высокий',
        'vacation_days_shortterm': 'Дни неотгуленного отпуска',
        'seniority': 'Стаж в компании',
        'income_vs_industry': 'Доход / средний доход по отрасли',
        'income_vs_position': 'Доход / средний доход по позиции',
        'income_per_experience': 'Доход / стаж',
        'position_median_income': 'Медианный доход на данной позиции',
        'income_shortterm': 'Средний доход (3 месяца)',
        'log_seniority': 'Логарифм стажа в компании',
        'age_group_36-45': 'Возрастная группа 36-45',
        'total_seniority': 'Общий стаж (трудовая книжка)',
        'absenteeism_shortterm': 'Отсутствия (3 месяца)',
        'absences_per_experience': 'Отсутствия / общий стаж',
        'absences_per_year': 'Отсутствия за год',
        'job_category_производство высококвалифицированное': 'Категория "Высококвалифицированное производство"',
        'vacations_by_city': 'Число актуальных вакансий в городе',
        'city_Невинномысск': 'Город: Невинномысск',
        'education_income_1_high': 'Высокий доход + высшее образование',
        'education_income_1_medium': 'Средний доход + высшее образование',
        'citizenship_gender_0_1': 'Женщина, российское гражданство',
        'gender_1': 'Женский пол',
        'city_population': 'Население города',
        'gender_age_0_18-25': 'Мужчина, 18-25 лет',
        'gender_age_0_26-35': 'Мужчина, 26-35 лет',
        'age_group_18-25': 'Возраст 18-25',
        'age_group_26-35': 'Возраст 26-35',
        'age_group_56-65': 'Возраст 56-65',
        'education_income_3_medium': 'Образование неизвестно, доход средний',
        'education_income_3_medium_low': 'Образование неизвестно, доход средне-низкий',
        'education_income_1_low': 'Низкий доход + высшее образование',
        'education_income_2_high': 'Высокий доход + сред. спец. образование',
        'education_income_2_low': 'Низкий доход + сред. спец. образование',
        'education_income_2_medium': 'Средний доход + сред. спец. образование',
        'education_income_3_low': 'Образование неизвестно, доход низкий',
        'citizenship_gender_0_0': 'Мужчина, российское гражданство',
        'gender_age_1_46-55': 'Женщина, 46-55 лет',
        'gender_age_1_18-25': 'Женщина, 18-25 лет',
        'income_group_high': 'Уровень дохода: высокий',
        'unused_vacation_per_experience': 'Неотгуленный отпуск / стаж',
        'children': 'число детей',
        'education_2': 'Среднее специальное образование',
        'education_1': 'Высшее образование'
    }

    # Функция для переименования фич
    def translate_feature_name(feature_name):
        return feature_name_mapping.get(feature_name, feature_name)

    # Вычисляем важность (средние SHAP значения)
    importance = shap_values[:, :-1].mean(axis=0)
    importance_abs = abs(shap_values[:, :-1]).mean(axis=0)

    # Сортировка для первого графика (знаковые значения)
    sorted_idx_sign = np.argsort(np.abs(importance))  # Индексы от наименьшего к наибольшему
    sorted_features_sign = [translate_feature_name(feature_names[i]) for i in sorted_idx_sign]
    sorted_shap_sign = [importance[i] for i in sorted_idx_sign]

    print("Feature Importance (signed values):")
    for f, i in zip(sorted_features_sign, sorted_shap_sign):
        print(f"{f}: {i:.6f}")

    # Создаем график для знаковых значений
    colors = ['red' if val > 0 else 'blue' for val in sorted_shap_sign]
    plt.figure(figsize=(12, 10))
    plt.barh(sorted_features_sign, [abs(s) for s in sorted_shap_sign], color=colors)
    plt.title('Важность признаков CatBoost (знаковые значения)', fontsize=14)
    plt.xlabel('Среднее абсолютное значение SHAP')
    plt.tight_layout()
    #plt.show()
    plt.clf()

    # Сортировка для второго графика (абсолютные значения) - топ 25
    sorted_idx_abs = np.argsort(importance_abs)[::-1]  # Индексы от наибольшего к наименьшему

    # Берем только топ 25
    top_25_idx = sorted_idx_abs[:40]
    top_25_features = [translate_feature_name(feature_names[i]) for i in top_25_idx]
    top_25_abs_shap = [importance_abs[i] for i in top_25_idx]

    # Создаем график для абсолютных значений (топ 25)
    plt.figure(figsize=(12, 10))
    plt.barh(top_25_features, top_25_abs_shap, color='skyblue')
    plt.gca().invert_yaxis()  # Чтобы самый важный признак был сверху
    plt.title('Топ 25 самых важных признаков (абсолютные значения)', fontsize=14)
    plt.xlabel('Среднее абсолютное значение SHAP')
    plt.tight_layout()
   # plt.show()
    plt.clf()

    #shap.dependence_plot("average_responses_count_9_historical", shap_values[:,:-1], _x_train)


    # Дополнительно: выводим топ 25 фичей
    print("\nТоп 25 самых важных признаков:")
    for i, (feature, importance_val) in enumerate(zip(top_25_features, top_25_abs_shap), 1):
        print(f"{i:2d}. {feature}: {importance_val:.6f}")

    return model


def save_value_to_csv(ba, r, p, filename='values_employees.csv'):
    """
    Append a floating point value to a CSV file in column 0

    Args:
        value: float value to save
        filename: name of the CSV file
    """
    # Create a DataFrame with the value
    new_row = pd.DataFrame({'ba': [ba], 'r': [r], 'p': [p]})

    # Append to existing file or create new one
    if os.path.exists(filename):
        # Read existing file and append new row
        existing_df = pd.read_csv(filename)
        updated_df = pd.concat([existing_df, new_row], ignore_index=True)
    else:
        # Create new file
        updated_df = new_row

    # Save to CSV
    updated_df.to_csv(filename, index=False)
    print(f"Values appended to {filename}")

def train(_x_train, _y_train, _x_test, _y_test, _sample_weight, _cat_feats_encoded, _model_name, _num_iters):
    if _model_name == 'XGBoostClassifier':
        model = train_xgboost_classifier(_x_train, _y_train, _x_test, _y_test, _sample_weight, _num_iters)
    elif _model_name == "CatBoostClassifier":
        model = train_catboost(_x_train, _y_train, _x_test, _y_test, _sample_weight, _cat_feats_encoded, _num_iters)
    else:
        print("Model name error: this model is not implemented yet!")
        return

    return model


def normalize(_data: pd.DataFrame):
    return _data.apply(lambda x: (x - x.min()) / (x.max() - x.min()), axis=0)


def create_shap_barchart(feature_names, shap_values, top_n=5):
    """Create a bar chart of top N SHAP values and return as image bytes"""
    # Get top N features by absolute SHAP value
    shap_series = pd.Series(shap_values, index=feature_names)
    top_shap = shap_series.abs().sort_values(ascending=False).head(top_n)

    # Create the bar chart
    fig, ax = plt.subplots(figsize=(6, 3))
    colors = ['red' if x < 0 else 'green' for x in shap_series[top_shap.index]]
    bars = ax.barh(range(len(top_shap)), shap_series[top_shap.index], color=colors)
    ax.set_yticks(range(len(top_shap)))
    ax.set_yticklabels([f"{name[:20]}..." if len(name) > 20 else name for name in top_shap.index])
    ax.set_xlabel('SHAP Value')
    ax.set_title('Top SHAP Features')
    plt.tight_layout()

    # Save to bytes buffer
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=80, bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)

    return buf, top_shap.index.tolist(), [shap_series[feature] for feature in top_shap.index]


def calc_empirical_threshold(proba_scores, recall, precision):
    # Параметры модели
    actual_attrition_rate = 0.113

    # Вычисляем целевую долю предсказаний
    target_positive_rate = (recall / precision) * actual_attrition_rate

    # Сортируем по убыванию вероятности
    sorted_proba = np.sort(proba_scores)[::-1]
    n_total = len(sorted_proba)

    # Находим порог
    n_target = int(target_positive_rate * n_total)
    threshold = sorted_proba[n_target - 1] if n_target > 0 else sorted_proba[0]

    print(f"Recall: {recall}, Precision: {precision}")
    print(f"Реальная текучесть: {actual_attrition_rate:.1%}")
    print(f"Ожидаемая доля положительных предсказаний: {target_positive_rate:.3f} ({target_positive_rate * 100:.1f}%)")
    print(f"Порог: {threshold:.4f}")
    print(f"Будет предсказано уходов: {n_target} из {n_total}")

    # Применяем порог
    predictions = (proba_scores >= threshold).astype(int)
    return threshold

def calc_threshold_by_hist(y_proba):
    prob_series = pd.Series(y_proba)

    # Create histogram
    plt.figure(figsize=(10, 6))
    ax = prob_series.hist(bins=np.arange(0, 1.05, 0.05),
                          edgecolor='black', alpha=0.7,
                          color='lightgreen', grid=False)

    plt.xlabel('Predicted Probability')
    plt.ylabel('Frequency')
    plt.title('CatBoost Prediction Probabilities Distribution')
    plt.xticks(np.arange(0, 1.1, 0.1))
    plt.grid(axis='y', alpha=0.3)
    #plt.show()
    plt.clf()
    return 0

def test(_model, _test_data, _cat_features, _inference=False):
    pos_class = 700
    neg_class = 5500

    test_data_all = pd.concat(_test_data, axis=0)
    trg = test_data_all['status']
    feat = test_data_all.drop(columns=['code'], errors='ignore')

    N_1 = len([y for y in trg if y == 1])
    N_0 = len([y for y in trg if y == 0])

    def adjusted_precision(P, N, M, new_N, new_M):
        numerator = P * (new_N / N)
        denominator = numerator + (1 - P) * (new_M / M)
        return numerator / denominator

    cat_columns = [col for col in feat.columns if col.startswith('job_category_2_')]
    for cat_col in cat_columns:
        print(f'\nCategory: {cat_col}')
        cat_name = cat_col.replace('job_category_2_', '')

        #if cat_name not in ['электромонтер, электрик', 'оператор, аппаратчик', 'составитель поездов, ремонтник вагонов, монтер', 'производство, разное']:
        #    continue

        # Extract the actual category name from the column name
        # Assuming format is 'job_category_2_actualname'

        for seniority in [100]:
            print(f'Seniority {seniority}')

            # Filter data where this category column == 1
            mask = (feat[cat_col] == 1) & (feat['seniority'] < seniority)

            if mask.sum() > 100:  # Check if there's data
                print(f'Amount: {mask.sum()}')
                # predictions = _model.predict_proba(feat[mask][feat['status']==1].drop(columns=['status']))
                # y_proba_united = predictions[:, 1]
                # calc_threshold_by_hist(y_proba_united)
                # predictions = _model.predict_proba(feat[mask][feat['status']==0].drop(columns=['status']))
                # y_proba_united = predictions[:, 1]
                # calc_threshold_by_hist(y_proba_united)

                predictions = _model.predict_proba(feat[mask].drop(columns=['status']))
                y_proba_united = predictions[:, 1]

                # Make sure target has matching indices
                trg_filtered = trg[mask]

                if len(np.unique(trg_filtered)) > 1:  # Need both classes for PR curve
                    precision, recall, thresholds = precision_recall_curve(trg_filtered, y_proba_united)

                f1_scores = 2 * (precision * recall) / (precision + recall + 1e-9)

                # Находим порог с максимальным F1
                optimal_idx = np.argmax(f1_scores)
                optimal_threshold = thresholds[optimal_idx]

                thresholds = [optimal_threshold]

                for threshold in thresholds:
                    print(f"Testing on united data with threshold = {threshold}...")
                    predictions_bin = (predictions[:, 1] > threshold).astype(int)

                    f1_united = f1_score(trg_filtered, predictions_bin)
                    recall_united = recall_score(trg_filtered, predictions_bin)
                    precision_united = precision_score(trg_filtered, predictions_bin)
                    #precision_united = adjusted_precision(precision_united, N_1, N_0, pos_class, neg_class)
                    ba = balanced_accuracy_score(trg_filtered, predictions_bin)

                    print(f"CatBoost result: F1 = {f1_united:.2f}, Recall = {recall_united:.2f}, Precision - {precision_united:.2f}, ba = {ba:.2f}, pos_pred = {predictions_bin.mean():.2f}%")
                save_value_to_csv(ba, recall_united, precision_united)



                result = confusion_matrix(trg_filtered, predictions_bin)
                sn.set(font_scale=1.4)  # for label size
                sn.heatmap(result, annot=True, annot_kws={"size": 16}, fmt='g')  # font size

                #plt.show()
                plt.clf()

                empirical_thresh=calc_empirical_threshold(y_proba_united, 45, 39)
                predictions_bin = (predictions[:, 1] > empirical_thresh).astype(int)

                f1_united = f1_score(trg_filtered, predictions_bin)
                recall_united = recall_score(trg_filtered, predictions_bin)
                precision_united = precision_score(trg_filtered, predictions_bin)
                precision_united = adjusted_precision(precision_united, N_1, N_0, pos_class, neg_class)
                ba = balanced_accuracy_score(trg_filtered, predictions_bin)

                print(f"CatBoost result with empirical threshold: F1 = {f1_united:.2f}, Recall = {recall_united:.2f}, Precision - {precision_united:.2f}, ba = {ba:.2f}")

    print("Testing separately...")
    for t in _test_data:
        trg = t['status']
        trn = t.drop(columns=['status', 'code'])
        N_1 = len([y for y in trg if y == 1])
        N_0 = len([y for y in trg if y == 0])

        # find opti thesh:
        predictions = _model.predict_proba(trn)
        y_proba = predictions[:, 1]
        precision, recall, thresholds = precision_recall_curve(trg, y_proba)

        f1_scores = 2 * (precision * recall) / (precision + recall + 1e-9)

        # Находим порог с максимальным F1
        optimal_idx = np.argmax(f1_scores)
        optimal_threshold = thresholds[optimal_idx]
        predictions = (predictions[:, 1] > optimal_threshold).astype(int)
        f1 = f1_score(trg, predictions)
        r = recall_score(trg, predictions)
        p = precision_score(trg, predictions)
        p = adjusted_precision(p, N_1, N_0, pos_class, neg_class)
        ba = balanced_accuracy_score(trg, predictions)

        print(f"test on {len(trg)} samples: thrs = {optimal_threshold}, F1={f1:.2f}, Recall={r:.2f}, Precision={p:.2f}, ba={ba:.2f}")

        predicted_classes = predictions  # (np.array(predictions) > 0.5).astype(int)
        result = confusion_matrix(trg, predicted_classes, )
        sn.set(font_scale=1.4)  # for label size
        sn.heatmap(result, annot=True, annot_kws={"size": 16}, fmt='g')  # font size

        #plt.show()
    plt.clf()

    def inference(__model, __dataset, __threshold):
        print('Apply rowwise...')
        # Create lists to store results
        results = []

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Predictions"

        # Write headers - including columns for top SHAP features
        headers = ['Code', 'Prediction Probability', 'Prediction Class']
        shap_headers = [f'Top_{i + 1}_Feature' for i in range(5)] + [f'Top_{i + 1}_SHAP' for i in range(5)]
        all_headers = headers + shap_headers + ['SHAP Chart']

        for col_idx, header in enumerate(all_headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row_idx = 2  # Start from row 2

        for n, row in __dataset.iterrows():
            if row['status'] == 0:
                code = row['code']
                # if code not in pd.read_excel('himik_predictions.xlsx')['Code'].values:
                #     continue
                term_date = 11

                # Create sample with the same features as batch processing
                sample = __dataset.loc[__dataset['code'] == code]
                feat = sample.drop(columns=['status', 'code'])

                # Get prediction probability
                pred_proba = __model.predict_proba(feat)[0, 1]

                # Apply the same threshold as batch processing
                prediction = 1 if pred_proba > __threshold else 0

                # Write basic data to Excel for ALL predictions
                ws.cell(row=row_idx, column=1, value=code)
                ws.cell(row=row_idx, column=2, value=pred_proba)
                ws.cell(row=row_idx, column=3, value=prediction)

                # Only create charts for predictions == 1
                # if prediction == 1:
                #     # Get SHAP values
                #     train_pool = Pool(data=feat, label=sample['status'])
                #     shap_values = __model.get_feature_importance(prettified=False, type='ShapValues', data=train_pool)
                #
                #     # Extract SHAP values (excluding base value)
                #     shap_for_prediction = shap_values[0, :-1]
                #     feature_names = feat.columns.tolist()
                #
                #     # Create SHAP bar chart and get top features
                #     chart_buffer, top_features, top_shap_values = create_shap_barchart(feature_names, shap_for_prediction,
                #                                                                        top_n=5)
                #
                #     # Write top SHAP features and values
                #     for i, (feature, shap_val) in enumerate(zip(top_features, top_shap_values), 1):
                #         ws.cell(row=row_idx, column=4 + i, value=feature)  # Feature name
                #         ws.cell(row=row_idx, column=4 + 5 + i, value=shap_val)  # SHAP value
                #
                #     # Insert SHAP bar chart
                #     img = Image(chart_buffer)
                #     img.anchor = f'{openpyxl.utils.get_column_letter(4 + 10 + 1)}{row_idx}'  # Column after SHAP values
                #     img.width = 300
                #     img.height = 150
                #     ws.add_image(img)
                # else:
                #     # For negative predictions, you might want to add some placeholder or explanation
                #     ws.cell(row=row_idx, column=5, value="No chart - prediction = 0")

                row_idx += 1  # Always increment row index

        # Adjust column widths
        for col in range(1, 20):  # Adjust first 19 columns
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Save the workbook
        output_file = 'himik_prediction_detailed_056.xlsx'
        wb.save(output_file)
        print(f"Detailed predictions with SHAP charts saved to {output_file}")

    if _inference:
        inference(_model, _test_data[0], 0.56)

    return f1_united, recall_united, precision_united

def calc_weights(_y_train: pd.DataFrame, _y_val: pd.DataFrame):
    w_0, w_1 = 0, 0
    print(_y_train)
    for i in _y_train.values:
        w_0 += 1 - i.item()
        w_1 += i.item()

    tot = w_0 + w_1
    w_0 = w_0 / tot
    w_1 = w_1 / tot
    print(f"weights:", w_0, w_1)
    #return np.array([w_0 if i == 1 else w_1 for i in _y_train.values])
    return np.array([1 if i == 1 else 1 for i in _y_train.values])

def remove_outliers(dfs):
    # Features to process
    features_to_trim = [
        'vacation_days_deriv',
        'vacation_days_shortterm',
        'absenteeism_deriv',
        'absenteeism_shortterm',
        'income_deriv',
        'income_shortterm',
        'total_seniority',
        'seniority',
        'salary_vs_city'
    ]
    new_datasets = []
    for df in dfs:
        for f in features_to_trim:
            if f not in df.columns:
                continue
            f = [f]
        # Step 1: Drop rows with NaN in features_to_trim (critical!)
            df = df.dropna(subset=f).copy()

        # Step 2: Skip features with zero variance (avoid NaN z-scores)
        # valid_features = [f for f in features_to_trim if df[f].var() > 0]
        # if not valid_features:
        #     new_datasets.append(df)  # No features left to filter
        #     continue

        # Step 3: Calculate Z-scores only on valid features
            z_scores = np.abs(stats.zscore(df[f]))
            filtered_rows = (z_scores < 3).all(axis=1)
            df_clean = df[filtered_rows].copy().reset_index(drop=True)  # Reset index here


        # Step 4: Verify no NaN in target
        assert df_clean['status'].isna().sum() == 0, \
            f"Target has {df_clean['status'].isna().sum()} NaN values after filtering!"

        new_datasets.append(df_clean)

        # Debug info
        print(f"Original rows: {len(df)}")
        print(f"Rows after outlier removal: {len(df_clean)}")
        print(f"Rows removed: {len(df) - len(df_clean)}")
    return new_datasets


def main(_config: dict):
    data_path = _config['dataset_src']
    test_path = _config['test_src']

    # Debug:
    # trn = pd.read_csv(data_path + '/chemi_dataset_24_jun.csv')
    # tst = pd.read_csv(test_path + '/chemi_dataset_24_dec.csv')

    #mask = (trn['status'] == 1) & (pd.to_datetime(trn['termination_date']).dt.date >= date(year=2024, month=9, day=1))
    #rows_to_transfer = trn[mask].copy()

    # Remove these rows from first DataFrame
    #trn = trn[~mask].copy()

    # Add them to second DataFrame
    #tst = pd.concat([tst, rows_to_transfer], ignore_index=True)

    #mask = (tst['status'] == 1) & (pd.to_datetime(tst['termination_date']).dt.date < date(year=2024, month=9, day=1))
    #rows_to_transfer = tst[mask].copy()

    # Remove these rows from first DataFrame
    #tst = tst[~mask].copy()

    # Add them to second DataFrame
    #trn = pd.concat([trn, rows_to_transfer], ignore_index=True)

    # If I want to remove extra snapshots for status==0:
    # trn = trn[~((trn['status'] == 0) & (trn['code'].str.contains('s1')))].copy()
    # tst = tst[~((tst['status'] == 0) & (tst['code'].str.contains('s1')))].copy()

    # If I want to separate personal codes in test and train:
    #status_zero_df1 = trn[trn['status'] == 0]
    #random_sample = status_zero_df1.sample(n=min(1000, len(status_zero_df1)),
#                                           random_state=42)  # Handles cases with <1000 rows
    #selected_codes = set(random_sample['code'])

    # trn = trn.drop(random_sample.index)
    # mask_to_remove = (tst['status'] == 0) & (~tst['code'].isin(selected_codes))
    # tst = tst[~mask_to_remove]  # Keep rows where the mask is False

    # trn.to_csv(data_path + '/chemi_dataset_24_jun.csv')
    # tst.to_csv(test_path + '/chemi_dataset_24_dec.csv')
    # debug end


    datasets = collect_datasets(data_path, _config['remove_small_period'], _config['remove_invalid_deriv'], _config['additional_filtering'])
    test_datasets = collect_datasets(test_path, _config['remove_small_period'], _config['remove_invalid_deriv'], _config['additional_filtering'])
    all_datasets = datasets + test_datasets

    # common = pd.concat(all_datasets)
    # new_test = common.sample(n=5000, random_state=43)
    # new_train = common[~common.index.isin(new_test.index)]
    # datasets = [new_train]
    # test_datasets = [new_test]
    # all_datasets = datasets + test_datasets
    # new_test.to_csv('new_test_chemy.csv')
    # new_train.to_csv('new_train_chemy.csv')

    if _config['random_split']:
        # take random split instead of splitting by years
        concat_dataset = pd.concat(datasets + test_datasets)
        trn_dataset, tst_dataset = train_test_split(concat_dataset, test_size=7000, random_state=43)
    else:
        trn_dataset = pd.concat(datasets)
        tst_dataset = pd.concat(test_datasets)

    print(len(tst_dataset[tst_dataset['status']==1]), ' 1s in TEST')
    datasets = [trn_dataset.reset_index(drop=True)]
    test_datasets = [tst_dataset.reset_index(drop=True)]
    # trn_dataset.to_csv('data/rand_train_chem.csv')
    # tst_dataset.to_csv('data/rand_test_chem.csv')

    rand_states = [4] # range(5)  # [777, 42, 6, 1370, 5087]
    score = [0, 0, 0]

    if _config['remove_short_service']:
        datasets = remove_short_service(datasets)
        test_datasets = remove_short_service(test_datasets)
        print(1, len(test_datasets[0]))

    if _config['remove_outliers']:
        datasets = [df.copy().reset_index(drop=True) for df in datasets]
        test_datasets = [df.copy().reset_index(drop=True) for df in test_datasets]
        datasets = remove_outliers(datasets)
        test_datasets = remove_outliers(test_datasets)
        print(2, len(test_datasets[0]))

    if _config['calculated_features']:
        datasets, new_cat_feat = create_features_for_datasets(datasets)
        test_datasets, _ = create_features_for_datasets(test_datasets)
        _config['cat_features'] += new_cat_feat
        print(f"New cat features: {_config['cat_features']}")


    for split_rand_state in rand_states:
        d_train, d_val, d_test, cat_feats_encoded = prepare_dataset_2(datasets,
                                                                      test_datasets,
                                                                      _config['make_synthetic'],
                                                                      _config['encode_categorical'],
                                                                      _config['use_selected_features'],
                                                                      _config['test_split'],
                                                                      _config['cat_features'],
                                                                      split_rand_state)

        if _config['smote']:
            d_train = minority_class_resample(d_train,cat_feats_encoded)
            #d_val = minority_class_resample(d_val, cat_feats_encoded)
            #d_test = minority_class_resample(d_test, cat_feats_encoded)

        x_train, y_train, x_val, y_val = get_united_dataset(d_train, d_val, d_test)
        # x_train, x_test, y_train, y_test = prepare_dataset(dataset, config['test_split'], config['normalize'])

        print(f"X train: {x_train.shape[0]}, x_val: {x_val.shape[0]}, y_train: {y_train.shape[0]}, y_val: {y_val.shape[0]}")
        sample_weight = calc_weights(y_train, y_val)
        merged = pd.merge(
            x_train[['age',  'seniority', 'income_shortterm']].reset_index(drop=True),
            x_val[['age',  'seniority', 'income_shortterm']].reset_index(drop=True),
            how='inner',
            indicator=False
        )
        print(f"\n\nNumber of duplicate rows: {len(merged)}")
        duplicate_indices = x_train.index.isin(merged.index)

        # Remove duplicates from x_train and y_train
        x_train = x_train[~duplicate_indices]
        y_train = y_train[~duplicate_indices]

        print(f"x_train shape after removing duplicates: {x_train.shape}")
        print(f"y_train shape after removing duplicates: {y_train.shape}")

        x_train.to_csv('x_train.csv')
        x_val.to_csv('x_val.csv')
        y_train.to_csv('y_train.csv')
        # print(sample_weight)
        #trained_model = train(x_train.drop(columns='code', errors='ignore'), y_train, x_val.drop(columns='code', errors='ignore'), y_val, sample_weight, cat_feats_encoded, _config['model'], _config['num_iters'])

        with open('model.pkl', 'rb') as file:
           trained_model = pickle.load(file)
        # print('\nRun test on TRAIN set...')
        # test(trained_model, d_train, cat_feats_encoded)
        print('\nRun test on TEST set...')
        # f1, r, p = test(trained_model, d_val, cat_feats_encoded, False)
        test_model(trained_model, x_train.drop(columns='code', errors='ignore'), y_train, d_val, cat_feats_encoded)


    with open('model.pkl', 'wb') as f:
       print("Saving model..")
       pickle.dump(trained_model, f)


if __name__ == '__main__':
    # config_path = 'config.json'  # config file is used to store some parameters of the dataset
    config = {
        'model': 'CatBoostClassifier',  # options: 'RandomForestRegressor', 'RandomForestRegressor_2','RandomForestClassifier', 'XGBoostClassifier', 'CatBoostClassifier'
        'test_split': 0.25,  # validation/training split proportion
        'normalize': False,  # normalize input values or not
        'remove_short_service': False,
        'remove_small_period': False,
        'remove_invalid_deriv': False,
        'num_iters': 8,  # number of fitting attempts
        'dataset_src': 'data/trn-4',  # 'data/chemistry_trn_full_merged_add_short_service_rnd',
        'test_src': 'data/tst-4',  #  'data/chemistry_tst_full_merged_add_short_service_rnd',
        'encode_categorical': True,
        'calculated_features': True,
        'use_selected_features': False,
        'remove_outliers': True,
        'make_synthetic': None,  # options: 'sdv', 'ydata', None
        'smote': False,  # perhaps not needed for catboost and in case if minority : majority > 0.5
        'random_split': False,
        'additional_filtering': True,
        'remove_staff_reduction': True,
        'cat_features': ['gender', 'season', 'month', 'citizenship', 'job_category_2', 'job_category', 'field',  'education', 'family_status']
    }

    main(config)

# 97d0ae93-9dfc-4c2a-9183-a0420a4d0771

